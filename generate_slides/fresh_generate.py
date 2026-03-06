"""Create FRESH NotebookLM notebooks with TEXT-ONLY sources and open in browser.
No Wikipedia URLs - just clean text content for reliable generation.
"""
import asyncio
import sys
import re
import openpyxl
from pathlib import Path
from playwright.async_api import async_playwright

PROFILE_DIR = Path.home() / ".notebooklm" / "browser_profile"
CP_PATH = Path("Courseware/TGS-2026061582 - ISO 22301-2019 Business Continuity Management System/Course Proposal/CP_TIPL_iso22301.xlsx")


def log(msg):
    print(msg, flush=True)


def parse_cp_xlsx(path):
    """Parse CP xlsx file and return raw text content."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    text_lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text_lines.append(f"## {sheet}")
        for row in ws.iter_rows(values_only=True):
            row_text = [str(cell) if cell is not None else "" for cell in row]
            if any(row_text):
                text_lines.append(" | ".join(row_text))
    wb.close()

    full_text = "\n".join(text_lines)

    # Trim to course content
    start_pat = re.compile(r"1\s*-\s*Course\s*Particulars", re.IGNORECASE)
    end_pat = re.compile(r"4\s*-\s*Declarations", re.IGNORECASE)
    start_m = start_pat.search(full_text)
    end_m = end_pat.search(full_text)
    if start_m and end_m and end_m.start() > start_m.start():
        full_text = full_text[start_m.start():end_m.start()].strip()

    return full_text


def build_lu1_content(raw_text):
    """Build LU1 source text: Get Started on ISO 22301 BCM (T1-T4)."""
    content = f"""# ISO 22301:2019 Business Continuity Management System
## LU1: Get Started on ISO 22301:2019 BUSINESS CONTINUITY MANAGEMENT

### Course Overview
This is a professional WSQ training course on ISO 22301:2019 Business Continuity Management System.

### Topics Covered in this Deck:
- T1: Cost-benefit analysis of ISO 22301 business continuity plans
- T2: Composition of ISO 22301 business continuity team
- T3: Functions and roles of ISO 22301 business continuity team
- T4: Resources in a ISO 22301 business continuity plan

### Slide Structure Requirements:
- INTRO PAGES (first deck of course):
  1. Cover slide: 'ISO 22301:2019 Business Continuity Management System', Version 1.0, Course Code, 'Trainer:'
  2. Digital Attendance
  3. About the Trainer
  4. Let's Know Each Other
  5. Ground Rules
  6-7. Lesson Plan
  8-9. Skills Framework
  10. Learning Outcomes
  11-12. Course Outline
  13. Final Assessment
  14. Briefing
  15. Criteria for Funding

- PER TOPIC: Section header + 8-12 knowledge slides + Activity/Lab slide
- Total target: ~60-80 slides
- Style: White background, Arial font, dark navy titles, slide numbers
- Every slide: two-column layout with professional images/diagrams on right

### Full Course Content:
{raw_text[:25000]}
"""
    return content


def build_lu2a_content(raw_text):
    """Build LU2-A source text: BCM Planning (T1-T2)."""
    content = f"""# ISO 22301:2019 Business Continuity Management System
## LU2-A: ISO 22301:2019 BUSINESS CONTINUITY MANAGEMENT Planning (T1-T2)

### Topics Covered in this Deck:
- T1: Techniques in the development of ISO 22301 business continuity plans (Scoping, BIA, Risk Assessment)
- T2: Components of business continuity plans (Recovery strategies, Plan structure, Documentation)

### Slide Structure Requirements:
- PER TOPIC: Section header + 8-12 knowledge slides + Activity/Lab slide
- Total target: ~30-40 slides
- Style: White background, Arial font, dark navy titles, slide numbers
- Every slide: two-column layout with professional images/diagrams on right
- Use EXACT topic titles from source material

### Full Course Content:
{raw_text[:25000]}
"""
    return content


async def main():
    # Parse CP
    log("Parsing CP file...")
    if not CP_PATH.exists():
        log(f"ERROR: CP file not found at {CP_PATH}")
        return
    raw_text = parse_cp_xlsx(CP_PATH)
    log(f"Parsed CP: {len(raw_text)} chars")

    # Create notebooks via API
    log("Creating fresh notebooks via API...")
    from notebooklm import NotebookLMClient

    client = await NotebookLMClient.from_storage()
    nb_ids = {}

    async with client:
        # Create LU1 notebook
        log("Creating LU1 notebook...")
        nb1 = await client.notebooks.create(
            "ISO 22301 BCM - LU1: Get Started (T1-T4) [FRESH]"
        )
        nb_ids["LU1"] = nb1.id
        log(f"  LU1 notebook: {nb1.id}")

        # Add text source for LU1
        lu1_text = build_lu1_content(raw_text)
        src1 = await client.sources.add_text(
            nb1.id,
            "LU1 Course Material - ISO 22301 BCM",
            lu1_text[:50000]
        )
        log(f"  LU1 source added: {src1.id} ({len(lu1_text)} chars)")

        # Create LU2-A notebook
        log("Creating LU2-A notebook...")
        nb2 = await client.notebooks.create(
            "ISO 22301 BCM - LU2-A: BCM Planning (T1-T2) [FRESH]"
        )
        nb_ids["LU2-A"] = nb2.id
        log(f"  LU2-A notebook: {nb2.id}")

        # Add text source for LU2-A
        lu2a_text = build_lu2a_content(raw_text)
        src2 = await client.sources.add_text(
            nb2.id,
            "LU2-A Course Material - ISO 22301 BCM Planning",
            lu2a_text[:50000]
        )
        log(f"  LU2-A source added: {src2.id} ({len(lu2a_text)} chars)")

        # Wait for sources to process
        log("Waiting for sources to process...")
        await asyncio.sleep(5)

    log("")
    log("=== Notebooks created! Opening in browser... ===")
    log("")

    # Open in browser
    pw = await async_playwright().start()
    ctx = await pw.chromium.launch_persistent_context(
        user_data_dir=str(PROFILE_DIR),
        headless=False,
        channel="chrome",
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-first-run",
            "--no-default-browser-check",
        ],
        ignore_default_args=["--enable-automation"],
    )

    page1 = ctx.pages[0] if ctx.pages else await ctx.new_page()
    url1 = f"https://notebooklm.google.com/notebook/{nb_ids['LU1']}"
    log(f"Opening LU1: {url1}")
    await page1.goto(url1, timeout=60000, wait_until="domcontentloaded")
    await asyncio.sleep(3)

    if "accounts.google.com" in page1.url:
        log("ERROR: Not logged in! Run quick_login.py first.")
        await ctx.close()
        await pw.stop()
        return

    log(f"LU1 loaded: {page1.url}")

    page2 = await ctx.new_page()
    url2 = f"https://notebooklm.google.com/notebook/{nb_ids['LU2-A']}"
    log(f"Opening LU2-A: {url2}")
    await page2.goto(url2, timeout=60000, wait_until="domcontentloaded")
    await asyncio.sleep(3)
    log(f"LU2-A loaded: {page2.url}")

    log("")
    log("=== BOTH FRESH NOTEBOOKS ARE OPEN ===")
    log("  -> Click 'Slide Deck' in each notebook to generate slides")
    log("  -> These have TEXT-ONLY sources (no Wikipedia URLs)")
    log("  -> Close browser window when done")
    log("")

    # Keep alive
    try:
        while True:
            await asyncio.sleep(5)
            try:
                _ = page1.url
            except Exception:
                break
    except KeyboardInterrupt:
        pass

    try:
        await ctx.close()
    except Exception:
        pass
    await pw.stop()
    log("Done!")


if __name__ == "__main__":
    asyncio.run(main())
